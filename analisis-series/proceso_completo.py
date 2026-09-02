# -*- coding: utf-8 -*-
"""
proceso_completo.py
-------------------
Un solo paso: junta los spoolers de emision, los cruza con el Excel de
Analisis de Series y deja el HTML interactivo en Descargas.

    C:\\Spoolers\\ASEG_01_2020.txt ... ASEG_12_2026.txt   (la emision)
    Descargas\\Analisis_de_Series.xlsx                    (el pivote)
                        |
                        v
    Descargas\\resumen_series.html   +   Descargas\\emision_cruzada.csv

No arma el parquet ni consolida los 84 archivos completos: de cada spooler
se queda SOLO con los registros cuya poliza o clave aparece en el Excel,
que son unos cuantos cientos. Por eso corre en minutos y no en horas.

USO (terminal):
    python proceso_completo.py
    python proceso_completo.py --spoolers "D:\\Spoolers" --anios 2020 2026
    python proceso_completo.py --columnas NOMBRE_ASEGURADO CALLE TELEFONO

USO (Jupyter): pega el bloque en una celda y en otra celda corre
    main()
Para cambiar algo, pasale los mismos argumentos por sys.argv antes:
    import sys; sys.argv = ["x", "--spoolers", r"D:\Spoolers"]; main()
"""

import os
import re
import csv
import sys
import glob
import json
import time
import argparse
import datetime
import unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("\nFalta openpyxl. Instalalo con:\n    pip install openpyxl\n")

# ============================ CONFIGURACION ============================

CARPETA_SPOOLERS = r"C:\Spoolers"
DESCARGAS        = os.path.join(os.path.expanduser("~"), "Downloads")
ANIO_MIN, ANIO_MAX = 2020, 2026

# De donde bajar la plantilla si no esta junto al script.
URL_PLANTILLA = ("https://raw.githubusercontent.com/maxavil/CV_MAX/"
                 "claude/interactive-policy-summary-html-uksnjh/"
                 "analisis-series/plantilla.html")

# Un registro nuevo abre con la poliza: 12 digitos y un "|". Segun el relleno
# puede traer espacios de por medio, asi que se prueban varios patrones y gana
# el que mas lineas explique en cada archivo.
PATRONES_INICIO = [
    re.compile(rb"^\d{12}\|"),
    re.compile(rb"^\s*\d{12}\s*\|"),
    re.compile(rb"^\s*\d{9,18}\s*\|"),
]
NOMBRE_MES = re.compile(r"ASEG_(\d{2})_(\d{4})", re.IGNORECASE)

# Columnas de la emision que se traen si no pides otras.
PATRONES_COLUMNAS = [
    re.compile(r"(NOMBRE|RAZON.?SOC)"),
    re.compile(r"(DIRECC|DOMICIL|CALLE|COLONIA|MUNICIP|ESTADO|CODIGO_?POSTAL|^CP$|_CP$)"),
    re.compile(r"(TELEF|CELUL|^TEL|_TEL|LADA)"),
    re.compile(r"(APODER|REPRESENT|LEGAL)"),
    re.compile(r"(INCISO|AGENTE|OFICINA|FECHA|PRIMA|RFC)"),
]

# Columnas del Excel: nombre interno -> como puede venir escrito.
COLUMNAS_EXCEL = {
    "SERIE":         ["serie", "vin", "numero de serie", "no serie"],
    "POLIZA":        ["poliza", "no poliza", "numero de poliza"],
    "INCISO":        ["inciso"],
    "DESDE":         ["desde", "fecha desde", "vigencia desde", "fecha de emision", "emision"],
    "HASTA":         ["hasta", "fecha hasta", "vigencia hasta"],
    "ANIO":          ["ano", "anio", "year", "ejercicio"],
    "CLAVE":         ["id asegurado", "clave asegurado", "clave de asegurado",
                      "id de asegurado", "cve asegurado", "clave"],
    "ASEGURADO":     ["asegurado", "nombre asegurado", "contratante"],
    "CLASIFICACION": ["clasificacion", "clasif"],
    "ID_AGENTE":     ["id agente", "clave agente", "cve agente"],
    "AGENTE":        ["agente", "nombre agente"],
    "OFICINA_ID":    ["oficina id", "id oficina", "clave oficina", "cve oficina"],
    "OFICINA":       ["oficina", "sucursal"],
}
ORDEN = list(COLUMNAS_EXCEL.keys())


# ============================ LLAVES ============================
# El Excel trae la poliza con 18 digitos y el spooler abre con 12: la llave
# son los primeros 12. La clave de asegurado se normaliza a 10 con ceros.

def llave_poliza(v):
    d = re.sub(r"\D+", "", str(v or ""))
    return "" if not d else (d[:12] if len(d) >= 12 else d.zfill(12))


def llave_clave(v):
    d = re.sub(r"\D+", "", str(v or ""))
    return "" if not d else (d[-10:] if len(d) >= 10 else d.zfill(10))


def sin_llave(v):
    return not v or not re.search(r"[1-9]", v)


def limpiar(s):
    """Quita relleno y los bytes de control que mete el host."""
    s = str(s or "").strip()
    return "".join(c for c in s if unicodedata.category(c)[0] != "C")


# ============================ EL EXCEL ============================

def clave_texto(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9 ]+", " ", s.lower()).strip()


def a_texto(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def anio_de(bruto):
    for campo in ("ANIO", "DESDE"):
        v = bruto.get(campo)
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.year
        s = a_texto(v)
        if not s:
            continue
        m = re.fullmatch(r"(\d{4})(?:\.0)?", s)
        if m:
            return int(m.group(1))
        m = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", s)
        if m:
            a = int(m.group(3))
            return a if a > 100 else (2000 + a if a < 70 else 1900 + a)
    return None


def fecha_larga(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    s = a_texto(v)
    m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2})", s)
    if m:
        a = int(m.group(3))
        a = 2000 + a if a < 70 else 1900 + a
        return "%02d/%02d/%d" % (int(m.group(1)), int(m.group(2)), a)
    return s


def leer_excel(ruta):
    """Devuelve {'columnas': ORDEN, 'filas': [...]} con una fila por poliza."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        n_enc, encabezado = None, None
        for i, fila in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            etiquetas = [clave_texto(c) for c in fila]
            if any(e == "serie" for e in etiquetas) and any(e.startswith("poliza") for e in etiquetas):
                n_enc, encabezado = i, list(fila)
                break
        if not n_enc:
            continue

        etiquetas = [clave_texto(c) for c in encabezado]
        mapa = {}
        for interno, alias in COLUMNAS_EXCEL.items():
            for j, e in enumerate(etiquetas):
                if e and j not in mapa.values() and e in alias:
                    mapa[interno] = j
                    break
            if interno in mapa:
                continue
            for j, e in enumerate(etiquetas):
                if e and j not in mapa.values() and any(e.startswith(a) or a.startswith(e) for a in alias):
                    mapa[interno] = j
                    break
        if "SERIE" not in mapa or "POLIZA" not in mapa:
            continue

        print("  hoja '%s', encabezado en la fila %d" % (nombre, n_enc))
        faltan = [c for c in ORDEN if c not in mapa]
        if faltan:
            print("  columnas que no encontre (van vacias): %s" % ", ".join(faltan))

        filas, sin_anio = [], 0
        for cruda in ws.iter_rows(min_row=n_enc + 1, values_only=True):
            bruto = {c: (cruda[j] if j < len(cruda) else None) for c, j in mapa.items()}
            if not a_texto(bruto.get("SERIE")) and not a_texto(bruto.get("POLIZA")):
                continue
            anio = anio_de(bruto)
            if not anio:
                sin_anio += 1
                continue
            filas.append([anio if c == "ANIO" else
                          fecha_larga(bruto.get(c)) if c in ("DESDE", "HASTA") else
                          a_texto(bruto.get(c)) for c in ORDEN])
        if sin_anio:
            print("  %d fila(s) sin anio de emision: se omiten" % sin_anio)
        return {"columnas": ORDEN, "filas": filas}

    sys.exit("No encontre en '%s' una hoja con columnas Serie y Poliza." % ruta)


# ============================ LOS SPOOLERS ============================

def nombres_columnas(crudos):
    """Igual que consolidar_aseg.py, con el sufijo _2 del nombre repetido."""
    vistos, out = {}, []
    for i, c in enumerate(crudos):
        base = re.sub(r"[^\w]+", "_", limpiar(c)).strip("_").upper() or ("COL_%d" % (i + 1))
        if base in vistos:
            vistos[base] += 1
            base = "%s_%d" % (base, vistos[base])
        else:
            vistos[base] = 1
        out.append(base)
    return out


def detectar_inicio(ruta, muestra=600):
    """Que patron abre un registro en ESTE archivo, o None si cada linea ya
    es un registro completo."""
    crudas = []
    with open(ruta, "rb") as fh:
        fh.readline()
        for cruda in fh:
            crudas.append(cruda.rstrip(b"\r\n"))
            if len(crudas) >= muestra:
                break
    if not crudas:
        return None
    mejor, mejor_tasa = None, 0.5
    for patron in PATRONES_INICIO:
        tasa = sum(1 for l in crudas if patron.match(l)) / len(crudas)
        if tasa > mejor_tasa:
            mejor, mejor_tasa = patron, tasa
    return mejor


def registros(ruta, patron):
    """(texto, es_cabecera, n_lineas). Reconstruye los registros partidos:
    toda linea que no abre registro se pega a la anterior."""
    with open(ruta, "rb") as fh:
        yield fh.readline().decode("latin-1").rstrip("\r\n"), True, 1
        buffer, partes = None, 0
        for cruda in fh:
            cruda = cruda.rstrip(b"\r\n")
            if patron is None:
                if cruda.strip():
                    yield cruda.decode("latin-1"), False, 1
                continue
            if patron.match(cruda):
                if buffer is not None:
                    yield buffer.decode("latin-1"), False, partes
                buffer, partes = cruda, 1
            elif buffer is None:
                if cruda.strip():
                    buffer, partes = cruda, 1
            else:
                buffer += cruda
                partes += 1
        if buffer is not None:
            yield buffer.decode("latin-1"), False, partes


def puntuar_columna(j, muestra, fn, llaves):
    return sum(1 for f in muestra if j < len(f) and f[j] and fn(f[j]) in llaves)


def adivinar_llaves(columnas, muestra, pol_ok, cla_ok):
    """La columna de poliza es la que mas valores mete dentro de las llaves
    del Excel; el nombre solo desempata."""
    jp = max(range(len(columnas)),
             key=lambda j: (puntuar_columna(j, muestra, llave_poliza, pol_ok),
                            bool(re.search(r"POLIZ", columnas[j]))))
    jc = max(range(len(columnas)),
             key=lambda j: (puntuar_columna(j, muestra, llave_clave, cla_ok),
                            bool(re.search(r"CLAVE|ASEGURAD", columnas[j]))))
    return jp, jc


def elegir_columnas(columnas, pedidas, llaves):
    disponibles = [c for c in columnas if c not in llaves]
    if pedidas:
        faltan = [c for c in pedidas if c not in disponibles]
        if faltan:
            print("  ! no existen en la emision y se ignoran: %s" % ", ".join(faltan))
        return [c for c in pedidas if c in disponibles]
    sel = [c for c in disponibles if any(p.search(c) for p in PATRONES_COLUMNAS)]
    return sel or disponibles


def recorrer_spoolers(carpeta, anios, pol_ok, cla_ok, pedidas, colapsar=True):
    """Un barrido por todos los ASEG_MM_AAAA.txt del rango.

    Con 84 spoolers la misma poliza reaparece cada mes con los mismos datos.
    Por omision se colapsa a UN registro por poliza + clave + anio + inciso,
    el del mes mas reciente, que es el que interesa para la celda de ese anio.
    Con colapsar=False se guardan todos los meses que traigan datos distintos.

    Devuelve las columnas exportadas y las filas
    [poliza, clave, anio, mes, ...datos].
    """
    archivos = []
    for ruta in sorted(glob.glob(os.path.join(carpeta, "ASEG_*.txt"))):
        m = NOMBRE_MES.search(os.path.basename(ruta))
        if not m:
            print("  ! %s no trae MM_AAAA en el nombre, se omite" % os.path.basename(ruta))
            continue
        mes, anio = int(m.group(1)), int(m.group(2))
        if anios[0] <= anio <= anios[1]:
            archivos.append((anio, mes, ruta))
    archivos.sort()                      # cronologico: el ultimo mes manda
    if not archivos:
        sys.exit("No encontre ASEG_MM_AAAA.txt de %d a %d en %s" % (anios[0], anios[1], carpeta))

    print("  %d archivo(s) entre %d y %d\n" % (len(archivos), anios[0], anios[1]))

    columnas = salida_cols = None
    jp = jc = k_inciso = None
    jsel = []
    mejores = {}                         # llave -> fila (modo colapsado)
    filas, vistos = [], set()            # modo completo
    leidas = rechazadas = reconstruidas = 0
    t0 = time.time()

    cuantas = lambda: len(mejores) if colapsar else len(filas)

    def fijar_columnas(muestra):
        """Con la primera muestra se deciden las llaves y las variables."""
        nonlocal jp, jc, salida_cols, jsel, k_inciso
        jp, jc = adivinar_llaves(columnas, muestra, pol_ok, cla_ok)
        salida_cols = elegir_columnas(columnas, pedidas, (columnas[jp], columnas[jc]))
        jsel = [columnas.index(c) for c in salida_cols]
        k_inciso = next((k for k, c in enumerate(salida_cols) if re.search(r"INCISO", c)), None)
        print("  llave poliza: %s | llave clave: %s" % (columnas[jp], columnas[jc]))
        print("  %d variables exportadas: %s%s\n"
              % (len(salida_cols), ", ".join(salida_cols[:6]),
                 ", ..." if len(salida_cols) > 6 else ""))

    for anio, mes, ruta in archivos:
        nombre = os.path.basename(ruta)
        mb = os.path.getsize(ruta) / 1024 ** 2
        patron = detectar_inicio(ruta)
        antes = cuantas()
        pendientes = []                  # registros previos a fijar las llaves

        def guarda(campos, anio=anio, mes=mes):
            p, c = llave_poliza(campos[jp]), llave_clave(campos[jc])
            if (sin_llave(p) or p not in pol_ok) and (sin_llave(c) or c not in cla_ok):
                return
            datos = [campos[j] for j in jsel]
            fila = [p, c, str(anio), "%02d" % mes] + datos
            if colapsar:
                mejores[(p, c, anio, datos[k_inciso] if k_inciso is not None else "")] = fila
            else:
                firma = (p, c, anio, tuple(datos))
                if firma not in vistos:
                    vistos.add(firma)
                    filas.append(fila)

        for texto, es_cabecera, partes in registros(ruta, patron):
            if es_cabecera:
                cols = nombres_columnas(texto.split("|"))
                if columnas is None:
                    columnas = cols
                elif cols != columnas:
                    print("    ! %s trae otro encabezado; uso el del primero" % nombre)
                continue

            campos = [limpiar(c) for c in texto.split("|")]
            leidas += 1
            if partes > 1:
                reconstruidas += 1
            if len(campos) != len(columnas):
                rechazadas += 1          # un "|" dentro del dato
                continue

            if jp is None:
                pendientes.append(campos)
                if len(pendientes) < 300:
                    continue
                fijar_columnas(pendientes)
                for previo in pendientes:
                    guarda(previo)
                pendientes = []
                continue
            guarda(campos)

        if jp is None and pendientes:    # archivo con menos de 300 registros
            fijar_columnas(pendientes)
            for previo in pendientes:
                guarda(previo)

        print("  %-22s %6.0f MB  ->  %4d nuevas  (%s acumuladas)"
              % (nombre, mb, cuantas() - antes, f"{cuantas():,}"), flush=True)

    if jp is None:
        sys.exit("Los spoolers no traen registros utilizables.")

    if colapsar:
        filas = sorted(mejores.values(), key=lambda f: (f[0], f[2], f[3]))
    print("\n  %s registros leidos | %s conservados | %s rechazados | %s reconstruidos | %.0f s"
          % (f"{leidas:,}", f"{len(filas):,}", f"{rechazadas:,}",
             f"{reconstruidas:,}", time.time() - t0))
    return (["ANIO_SPOOLER", "MES_SPOOLER"] + salida_cols, filas,
            columnas[jp], columnas[jc], leidas, rechazadas)


# ============================ DIAGNOSTICO ============================

def resumen_cruce(datos, filas):
    """Cuantas polizas del Excel encuentran registro en la emision. La clave
    0000000000 no es llave, igual que en el HTML: no cuenta como match."""
    por_pol = {f[0] for f in filas if not sin_llave(f[0])}
    por_cla = {f[1] for f in filas if not sin_llave(f[1])}
    ip, ic = ORDEN.index("POLIZA"), ORDEN.index("CLAVE")

    n = cp = cc = ca = sin = 0
    for f in datos["filas"]:
        p, c = llave_poliza(f[ip]), llave_clave(f[ic])
        hp = not sin_llave(p) and p in por_pol
        hc = not sin_llave(c) and c in por_cla
        n += 1
        cp += hp
        cc += hc
        ca += hp and hc
        sin += not (hp or hc)

    pc = lambda x: "%5.1f%%" % (100.0 * x / n) if n else "  0.0%"
    print("\n  De las %d polizas del analisis:" % n)
    print("    con match por poliza : %4d  %s" % (cp, pc(cp)))
    print("    con match por clave  : %4d  %s" % (cc, pc(cc)))
    print("    con match por ambas  : %4d  %s" % (ca, pc(ca)))
    print("    sin match            : %4d  %s" % (sin, pc(sin)))
    if sin:
        print("    (el detalle poliza por poliza esta en 'Ver matches' del HTML)")


# ============================ PLANTILLA ============================

def obtener_plantilla(ruta_dada=None):
    candidatas = [ruta_dada] if ruta_dada else []
    aqui = os.path.dirname(os.path.abspath(__file__))
    candidatas += [os.path.join(aqui, "plantilla.html"),
                   os.path.join(aqui, "analisis-series", "plantilla.html"),
                   os.path.join(DESCARGAS, "plantilla.html")]
    for c in candidatas:
        if c and os.path.exists(c):
            print("Plantilla: %s" % c)
            return open(c, "r", encoding="utf-8").read()

    print("Plantilla: no esta en disco, la bajo del repositorio...")
    try:
        from urllib.request import urlopen
        html = urlopen(URL_PLANTILLA, timeout=30).read().decode("utf-8")
    except Exception as e:
        sys.exit("\nNo pude bajar la plantilla (%s).\nDeja plantilla.html junto a "
                 "este script o pasala con --plantilla.\n" % e)
    destino = os.path.join(aqui, "plantilla.html")
    try:
        open(destino, "w", encoding="utf-8").write(html)
        print("  guardada en %s" % destino)
    except OSError:
        pass
    return html


# ============================ PRINCIPAL ============================

def main():
    ap = argparse.ArgumentParser(description="Spoolers + Excel -> HTML interactivo")
    ap.add_argument("--spoolers", default=CARPETA_SPOOLERS)
    ap.add_argument("--excel", default=None)
    ap.add_argument("--salida", default=None)
    ap.add_argument("--csv", default=None, help="CSV de la emision cruzada")
    ap.add_argument("--anios", nargs=2, type=int, default=[ANIO_MIN, ANIO_MAX],
                    metavar=("DESDE", "HASTA"))
    ap.add_argument("--columnas", nargs="*", default=None,
                    help="variables de la emision a traer (por omision, las utiles)")
    ap.add_argument("--plantilla", default=None)
    ap.add_argument("--todos-los-meses", action="store_true",
                    help="guarda cada mes por separado en vez de uno por anio")
    # parse_known_args y no parse_args: en Jupyter el kernel mete su propio
    # "-f kernel-xxxx.json" en sys.argv y argparse aborta con "unrecognized
    # arguments". Asi el bloque corre igual en notebook que en la terminal.
    args, _ = ap.parse_known_args()

    base = DESCARGAS if os.path.isdir(DESCARGAS) else os.getcwd()
    excel = args.excel
    if not excel:
        for patron in ("*nalisis*eries*.xls*", "*Analisis*Series*.xls*"):
            hallados = sorted(glob.glob(os.path.join(base, patron)))
            if hallados:
                excel = hallados[0]
                break
    if not excel or not os.path.exists(excel):
        sys.exit("No encontre el Excel en %s. Pasalo con --excel \"ruta\\archivo.xlsx\"" % base)

    salida = args.salida or os.path.join(base, "resumen_series.html")
    ruta_csv = args.csv or os.path.join(base, "emision_cruzada.csv")

    print("=" * 66)
    print("  1/4  El pivote")
    print("=" * 66)
    print("Excel: %s" % excel)
    datos = leer_excel(excel)
    ip, ic = ORDEN.index("POLIZA"), ORDEN.index("CLAVE")
    pol_ok = {llave_poliza(f[ip]) for f in datos["filas"]}
    cla_ok = {llave_clave(f[ic]) for f in datos["filas"]}
    pol_ok = {v for v in pol_ok if not sin_llave(v)}
    cla_ok = {v for v in cla_ok if not sin_llave(v)}
    anios_excel = sorted({f[ORDEN.index("ANIO")] for f in datos["filas"]})
    print("  %d polizas, %d series, anios %d-%d"
          % (len(datos["filas"]), len({f[0] for f in datos["filas"]}),
             anios_excel[0], anios_excel[-1]))
    print("  a buscar en la emision: %d polizas y %d claves" % (len(pol_ok), len(cla_ok)))

    print("\n" + "=" * 66)
    print("  2/4  La emision")
    print("=" * 66)
    print("Carpeta: %s" % args.spoolers)
    cols, filas, col_p, col_c, leidas, rech = recorrer_spoolers(
        args.spoolers, args.anios, pol_ok, cla_ok, args.columnas,
        colapsar=not args.todos_los_meses)
    resumen_cruce(datos, filas)

    print("\n" + "=" * 66)
    print("  3/4  El CSV de respaldo")
    print("=" * 66)
    with open(ruta_csv, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["POLIZA12", "CLAVE_ASEGURADO"] + cols)
        w.writerows(filas)
    print("%s  (%.1f KB)" % (ruta_csv, os.path.getsize(ruta_csv) / 1024))

    print("\n" + "=" * 66)
    print("  4/4  El HTML")
    print("=" * 66)
    html = obtener_plantilla(args.plantilla)
    datos["generado"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    datos["fuente"] = os.path.basename(excel)
    emision = {
        "archivo": "emision %d-%d (%d spoolers)" % (args.anios[0], args.anios[1],
                                                    len(glob.glob(os.path.join(args.spoolers, "ASEG_*.txt")))),
        "colPoliza": col_p, "colClave": col_c,
        "cols": cols, "filas": filas,
    }

    def inyectar(html, marca, valor):
        a = "/*__%s__*/" % marca
        b = "/*__FIN_%s__*/" % marca
        return (html[:html.index(a)] +
                json.dumps(valor, ensure_ascii=False, separators=(",", ":")) +
                html[html.index(b) + len(b):])

    html = inyectar(html, "DATOS", datos)
    html = inyectar(html, "EMISION", emision)
    with open(salida, "w", encoding="utf-8") as fh:
        fh.write(html)

    print("%s  (%.1f MB)" % (salida, os.path.getsize(salida) / 1024 ** 2))
    print("\nListo. Abrelo con doble clic.")


if __name__ == "__main__":
    main()
