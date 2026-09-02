# -*- coding: utf-8 -*-
"""
construir_resumen.py
--------------------
Toma el Excel de Analisis de Series (SISE) y escupe un HTML interactivo:
el pivote a la izquierda, un anio por columna a la derecha, y botones para
cambiar el dato que se pinta en cada celda.

La base de emision NO se necesita aqui: el HTML la lee despues en el navegador
(arrastrando el spooler ASEG_MM_AAAA.txt o un CSV). Si quieres dejarla ya
incrustada, pasa --emision con el CSV que genera preparar_emision.py.

USO:
    python construir_resumen.py
    python construir_resumen.py --excel "C:\\Users\\max\\Downloads\\Analisis_de_Series.xlsx"
    python construir_resumen.py --salida "C:\\Users\\max\\Downloads\\resumen_series.html"
    python construir_resumen.py --emision "C:\\Spoolers\\consolidado\\emision_para_html.csv"
"""

import os
import re
import csv
import sys
import json
import glob
import argparse
import datetime
import unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("\nFalta openpyxl. Instalalo con:\n    pip install openpyxl\n")

AQUI = os.path.dirname(os.path.abspath(__file__))
PLANTILLA = os.path.join(AQUI, "plantilla.html")

# Nombre interno -> como puede venir escrito en el Excel.
COLUMNAS = {
    "SERIE":         ["serie", "vin", "numero de serie", "no serie"],
    "POLIZA":        ["poliza", "no poliza", "numero de poliza"],
    "INCISO":        ["inciso"],
    "DESDE":         ["desde", "fecha desde", "vigencia desde", "fecha de emision", "emision"],
    "HASTA":         ["hasta", "fecha hasta", "vigencia hasta"],
    "ANIO":          ["ano", "anio", "year", "ejercicio"],
    "CLAVE":         ["id asegurado", "clave asegurado", "clave de asegurado",
                      "id de asegurado", "cve asegurado", "clave"],
    "ASEGURADO":     ["asegurado", "nombre asegurado", "contratante"],
    "CLASIFICACION": ["clasificacion", "clasif"],
    "ID_AGENTE":     ["id agente", "clave agente", "cve agente"],
    "AGENTE":        ["agente", "nombre agente"],
    "OFICINA_ID":    ["oficina id", "id oficina", "clave oficina", "cve oficina"],
    "OFICINA":       ["oficina", "sucursal"],
}
ORDEN = list(COLUMNAS.keys())


def clave_texto(s):
    """minusculas, sin acentos, sin signos: para comparar encabezados."""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9 ]+", " ", s.lower()).strip()


def a_texto(v):
    """Celda -> texto plano, sin que openpyxl invente notacion cientifica."""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, datetime.date):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def anio_de(fila):
    """Anio de emision: primero la columna Anio, si no de la fecha Desde."""
    for campo in ("ANIO", "DESDE"):
        v = fila.get(campo)
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.year
        s = a_texto(v)
        if not s:
            continue
        m = re.fullmatch(r"(\d{4})(?:\.0)?", s)
        if m:
            return int(m.group(1))
        m = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", s)
        if m:
            a = int(m.group(3))
            return a if a > 100 else (2000 + a if a < 70 else 1900 + a)
    return None


def fecha_larga(v):
    """dd/mm/aa -> dd/mm/aaaa; lo demas se deja como viene."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    s = a_texto(v)
    m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2})", s)
    if m:
        a = int(m.group(3))
        a = 2000 + a if a < 70 else 1900 + a
        return "%02d/%02d/%d" % (int(m.group(1)), int(m.group(2)), a)
    return s


def localizar_encabezado(ws, max_filas=15):
    """La fila de encabezado es la primera que reconoce Serie y Poliza."""
    for i, fila in enumerate(ws.iter_rows(min_row=1, max_row=max_filas, values_only=True), 1):
        etiquetas = [clave_texto(c) for c in fila]
        if any(e == "serie" for e in etiquetas) and any(e.startswith("poliza") for e in etiquetas):
            return i, list(fila)
    return None, None


def mapear(encabezado):
    """{nombre interno: indice de columna} a partir del encabezado del Excel."""
    etiquetas = [clave_texto(c) for c in encabezado]
    mapa = {}
    for interno, alias in COLUMNAS.items():
        for j, e in enumerate(etiquetas):
            if not e or j in mapa.values():
                continue
            if e in alias:
                mapa[interno] = j
                break
        if interno in mapa:
            continue
        for j, e in enumerate(etiquetas):          # segunda pasada: por prefijo
            if not e or j in mapa.values():
                continue
            if any(e.startswith(a) or a.startswith(e) for a in alias):
                mapa[interno] = j
                break
    return mapa


def leer_excel(ruta, hoja=None):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hojas = [hoja] if hoja else wb.sheetnames
    for nombre in hojas:
        if nombre not in wb.sheetnames:
            continue
        ws = wb[nombre]
        n_enc, encabezado = localizar_encabezado(ws)
        if not n_enc:
            continue
        mapa = mapear(encabezado)
        faltan = [c for c in ("SERIE", "POLIZA") if c not in mapa]
        if faltan:
            continue

        print("  hoja '%s': encabezado en la fila %d" % (nombre, n_enc))
        no_vistas = [c for c in ORDEN if c not in mapa]
        if no_vistas:
            print("  columnas que no encontre (van vacias): %s" % ", ".join(no_vistas))

        filas, saltadas = [], 0
        for cruda in ws.iter_rows(min_row=n_enc + 1, values_only=True):
            bruto = {c: (cruda[j] if j < len(cruda) else None) for c, j in mapa.items()}
            if not a_texto(bruto.get("SERIE")) and not a_texto(bruto.get("POLIZA")):
                continue
            anio = anio_de(bruto)
            if not anio:
                saltadas += 1
                continue
            reg = []
            for c in ORDEN:
                if c == "ANIO":
                    reg.append(anio)
                elif c in ("DESDE", "HASTA"):
                    reg.append(fecha_larga(bruto.get(c)))
                else:
                    reg.append(a_texto(bruto.get(c)))
            filas.append(reg)
        if saltadas:
            print("  %d fila(s) sin anio de emision: se omiten" % saltadas)
        return {"hoja": nombre, "columnas": ORDEN, "filas": filas}

    sys.exit("No encontre en '%s' una hoja con columnas Serie y Poliza." % ruta)


def leer_emision_csv(ruta):
    """CSV de preparar_emision.py: col 1 poliza, col 2 clave, el resto datos."""
    with open(ruta, "r", encoding="utf-8-sig", newline="") as fh:
        r = csv.reader(fh)
        cab = next(r)
        filas = [f for f in r if len(f) == len(cab)]
    return {
        "archivo": os.path.basename(ruta),
        "colPoliza": cab[0], "colClave": cab[1],
        "cols": cab[2:],
        "filas": filas,
    }


def buscar_excel(pistas):
    for p in pistas:
        for ruta in sorted(glob.glob(p)):
            return ruta
    return None


def main():
    descargas = os.path.join(os.path.expanduser("~"), "Downloads")
    ap = argparse.ArgumentParser(description="Genera el HTML interactivo del analisis de series")
    ap.add_argument("--excel", default=None, help="Analisis_de_Series.xlsx")
    ap.add_argument("--hoja", default=None, help="hoja del Excel (por omision se busca sola)")
    ap.add_argument("--salida", default=None, help="HTML de salida (por omision Descargas)")
    ap.add_argument("--emision", default=None, help="CSV de preparar_emision.py para incrustar")
    args = ap.parse_args()

    excel = args.excel or buscar_excel([
        os.path.join(descargas, "*nalisis*eries*.xls*"),
        os.path.join(descargas, "*Analisis*Series*.xls*"),
        os.path.join(AQUI, "*.xlsx"),
    ])
    if not excel or not os.path.exists(excel):
        sys.exit("No encontre el Excel. Pasalo con --excel \"ruta\\Analisis_de_Series.xlsx\"")

    salida = args.salida or os.path.join(
        descargas if os.path.isdir(descargas) else AQUI, "resumen_series.html")

    print("Excel   : %s" % excel)
    datos = leer_excel(excel, args.hoja)
    datos["generado"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    datos["fuente"] = os.path.basename(excel)

    anios = sorted({f[ORDEN.index("ANIO")] for f in datos["filas"]})
    print("  %d registros, %d series, anios %s-%s"
          % (len(datos["filas"]),
             len({f[0] for f in datos["filas"]}),
             anios[0] if anios else "?", anios[-1] if anios else "?"))

    emision = None
    if args.emision:
        emision = leer_emision_csv(args.emision)
        print("Emision : %s (%d filas, %d variables)"
              % (args.emision, len(emision["filas"]), len(emision["cols"])))

    with open(PLANTILLA, "r", encoding="utf-8") as fh:
        html = fh.read()

    def inyectar(html, marca, valor):
        ini = html.index("/*__%s__*/" % marca)
        fin = html.index("/*__FIN_%s__*/" % marca) + len("/*__FIN_%s__*/" % marca)
        return html[:ini] + json.dumps(valor, ensure_ascii=False, separators=(",", ":")) + html[fin:]

    html = inyectar(html, "DATOS", datos)
    html = inyectar(html, "EMISION", emision)

    os.makedirs(os.path.dirname(os.path.abspath(salida)) or ".", exist_ok=True)
    with open(salida, "w", encoding="utf-8") as fh:
        fh.write(html)

    print("\nListo -> %s  (%.1f KB)" % (salida, os.path.getsize(salida) / 1024))
    print("Abrelo con doble clic; la emision se carga desde el propio HTML.")


if __name__ == "__main__":
    main()
