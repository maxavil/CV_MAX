#!/usr/bin/env python3
"""Lectura de los dos Excel de reservas técnicas QES y armado del histórico.

Equivalente en Python del bloque de ingesta de reservas/index.html.

    pip install openpyxl pyxlsb
    python reservas_qes.py Balanza_062026.xlsx ResultadosQES.xlsb

Reglas:
  * Metodología local  -> balanza de comprobación, columna de 3er grado de las
    cuentas 2205, 2301 y 2302. Los saldos vienen en negativo por ser pasivos y
    se les invierte el signo.
  * Método Estatutario CNSF -> archivo de actuarios, concepto por concepto,
    tomando todos los cortes que traiga el archivo.
  * Diferencia = estatutario - local.

Cada archivo procesado actualiza su periodo y conserva los demás: el histórico
se guarda en un JSON y se va conversando con el mes que sigue.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Sequence

# =========================================================================
# 1. Configuración de conceptos
# =========================================================================


@dataclass(frozen=True)
class Concepto:
    id: str
    label: str
    cuenta: str
    patron: re.Pattern


CONCEPTOS: list[Concepto] = [
    Concepto("rrc",  "Reserva de Riesgos en Curso",         "2205", re.compile(r"riesgos en curso")),
    Concepto("rsr",  "Reserva de Siniestros Reportados",    "2301", re.compile(r"siniestros reportados")),
    Concepto("rsnr", "Reserva de Siniestros No Reportados", "2302", re.compile(r"no reportados")),
]

GRADO = 3  # columna de la balanza de la que se toma el saldo

MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
         "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

# =========================================================================
# 2. Utilidades
# =========================================================================


def norm(v: Any) -> str:
    """Minúsculas, sin acentos y con espacios colapsados."""
    s = "" if v is None else str(v)
    s = unicodedata.normalize("NFD", s.lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip()


def to_num(v: Any) -> float | None:
    """Número de una celda; acepta texto con comas, signo $ y paréntesis."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return None
    s = re.sub(r"[\s$]", "", v).replace(",", "")
    neg = bool(re.fullmatch(r"\(.*\)", s))
    if neg:
        s = s[1:-1]
    if not re.fullmatch(r"-?\d*\.?\d+", s):
        return None
    n = float(s)
    return -n if neg else n


def ultimo_dia(anio: int, mes: int) -> int:
    siguiente = dt.date(anio + (mes == 12), (mes % 12) + 1, 1)
    return (siguiente - dt.timedelta(days=1)).day


def serial_a_fecha(n: float) -> str | None:
    """Serial de Excel -> 'AAAA-MM-DD' (base 1899-12-30)."""
    if not isinstance(n, (int, float)) or isinstance(n, bool):
        return None
    if n < 20000 or n > 80000:
        return None
    d = dt.date(1899, 12, 30) + dt.timedelta(days=int(round(n)))
    return d.isoformat()


def celda_a_fecha(v: Any) -> str | None:
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return serial_a_fecha(v)
    return None


def parse_fecha(texto: Any) -> str | None:
    """'al 30 de Junio 2026' / '30 de junio de 2026' -> '2026-06-30'."""
    m = re.search(r"(\d{1,2})\s+de\s+([a-z]+)\s*(?:de\s*)?(\d{4})", norm(texto))
    if not m:
        return None
    if m.group(2) not in MESES:
        return None
    return dt.date(int(m.group(3)), MESES.index(m.group(2)) + 1, int(m.group(1))).isoformat()


def periodo_de_nombre(nombre: str) -> str | None:
    """'Balanza_062026.xlsx' -> '2026-06-30'."""
    m = re.search(r"(0[1-9]|1[0-2])[_\-.]?(20\d{2})", str(nombre))
    if not m:
        return None
    mes, anio = int(m.group(1)), int(m.group(2))
    return dt.date(anio, mes, ultimo_dia(anio, mes)).isoformat()


def etiqueta_periodo(k: str) -> str:
    a, m, d = k.split("-")
    return f"{int(d)} {MESES[int(m) - 1]} {a}"


def etiqueta_corta(k: str) -> str:
    a, m, _ = k.split("-")
    return f"{MESES[int(m) - 1].capitalize()} {a}"


def match_concepto(texto: Any) -> Concepto | None:
    s = norm(texto)
    if "reserva" not in s:
        return None
    for c in CONCEPTOS:
        if c.id == "rsr" and "no reportados" in s:
            continue
        if c.patron.search(s):
            return c
    return None


def fmt(v: float) -> str:
    return f"{v:,.2f}"


# =========================================================================
# 3. Apertura del libro (.xlsx / .xlsm y .xlsb con el mismo resultado)
# =========================================================================

Hoja = tuple[str, list[list[Any]]]


def leer_libro(ruta: str | Path) -> list[Hoja]:
    """Devuelve [(nombre de hoja, filas como listas densas)] para xlsx o xlsb."""
    ruta = Path(ruta)
    ext = ruta.suffix.lower()

    if ext == ".xlsb":
        try:
            from pyxlsb import open_workbook
        except ImportError as err:  # pragma: no cover
            raise SystemExit("Para leer .xlsb hace falta pyxlsb: pip install pyxlsb") from err
        hojas: list[Hoja] = []
        with open_workbook(str(ruta)) as wb:
            for nombre in wb.sheets:
                with wb.get_sheet(nombre) as sh:
                    filas: list[list[Any]] = []
                    for fila in sh.rows():
                        densa: list[Any] = []
                        for celda in fila:
                            while len(densa) < celda.c:
                                densa.append(None)
                            densa.append(celda.v)
                        filas.append(densa)
                    hojas.append((nombre, filas))
        return hojas

    if ext in (".xlsx", ".xlsm", ".xltx"):
        try:
            import openpyxl
        except ImportError as err:  # pragma: no cover
            raise SystemExit("Para leer .xlsx hace falta openpyxl: pip install openpyxl") from err
        wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
        hojas = [(ws.title, [list(f) for f in ws.iter_rows(values_only=True)]) for ws in wb.worksheets]
        wb.close()
        return hojas

    raise SystemExit(f"Extensión no soportada: {ruta.name} (usa .xlsx, .xlsm o .xlsb)")


# =========================================================================
# 4. Lectura de la balanza de comprobación
# =========================================================================


@dataclass
class LecturaBalanza:
    periodo: str | None
    valores: dict[str, float]
    detalle: list[str]
    faltantes: list[str]
    hoja: str


def parse_balanza(hojas: Sequence[Hoja], fname: str) -> LecturaBalanza | None:
    """Metodología local: cuentas 2205 / 2301 / 2302 en la columna de 3er grado."""
    for nombre, filas in hojas:
        h = next((i for i, f in enumerate(filas[:60]) if f and norm(f[0]) == "cuenta"), None)
        if h is None:
            continue  # sin encabezado CUENTA no es una balanza

        # columnas por grado: la fila de encabezado dice "GRADO" y la de arriba el ordinal
        grados: dict[int, int] = {}
        hdr = filas[h]
        arriba = filas[h - 1] if h > 0 else []
        for j in range(1, len(hdr)):
            if "grado" in norm(hdr[j]):
                m = re.search(r"(\d)", str(arriba[j]) if j < len(arriba) and arriba[j] is not None else "")
                if m:
                    grados[int(m.group(1))] = j
        cols_grado = list(grados.values()) or list(range(2, 9))

        # índice de cuentas
        cuentas: dict[str, list[Any]] = {}
        for fila in filas[h + 1:]:
            if not fila or fila[0] is None:
                continue
            clave = str(fila[0]).strip()
            if clave and clave not in cuentas:
                cuentas[clave] = fila

        # periodo: encabezado del reporte, si no, nombre del archivo
        periodo = None
        for fila in filas[:h + 1]:
            periodo = parse_fecha(" ".join("" if c is None else str(c) for c in (fila or [])))
            if periodo:
                break
        if not periodo:
            periodo = periodo_de_nombre(fname)

        valores: dict[str, float] = {}
        detalle: list[str] = []
        faltantes: list[str] = []
        for c in CONCEPTOS:
            fila = cuentas.get(c.cuenta)
            if fila is None:
                faltantes.append(c.cuenta)
                continue
            val = col_usada = None
            j = grados.get(GRADO)
            if j is not None and j < len(fila):
                val = to_num(fila[j])
                if val is not None:
                    col_usada = GRADO
            if val is None:  # respaldo: primer grado con importe en esa fila
                for col in cols_grado:
                    if col >= len(fila):
                        continue
                    v = to_num(fila[col])
                    if v is not None:
                        val = v
                        col_usada = next((g for g, cc in grados.items() if cc == col), None)
                        break
            if val is None:
                faltantes.append(c.cuenta)
                continue
            valores[c.id] = abs(val)  # pasivo en negativo -> se invierte
            grado_txt = f"{col_usada}º grado " if col_usada else ""
            detalle.append(f"{c.cuenta} {grado_txt}{fmt(abs(val))}")

        if not valores:
            continue
        return LecturaBalanza(periodo, valores, detalle, faltantes, nombre)

    return None  # no es balanza -> se intenta como archivo de actuarios


# =========================================================================
# 5. Lectura del archivo de actuarios (todas las fechas que traiga)
# =========================================================================


@dataclass
class LecturaActuarios:
    periodos: dict[str, dict[str, dict[str, float]]] = field(default_factory=dict)
    hojas: list[str] = field(default_factory=list)


def parse_actuarios(hojas: Sequence[Hoja]) -> LecturaActuarios:
    """Busca el texto de cada concepto y toma los dos primeros importes a su derecha."""
    out = LecturaActuarios()
    for nombre, filas in hojas:
        titulo: str | None = None
        usada = False
        for fila in filas:
            if not fila:
                continue
            concepto: Concepto | None = None
            ci = -1
            for j, v in enumerate(fila):
                if not isinstance(v, str):
                    continue
                c = match_concepto(v)
                if c and ci < 0:
                    concepto, ci = c, j
                elif not c:
                    f = parse_fecha(v)  # título del bloque: "30 de junio de 2026"
                    if f:
                        titulo = f
            if concepto is None:
                continue

            periodo = next((p for p in (celda_a_fecha(fila[j]) for j in range(ci)) if p), None) or titulo
            if not periodo:
                continue

            nums: list[float] = []
            for j in range(ci + 1, len(fila)):
                n = to_num(fila[j])
                if n is not None:
                    nums.append(abs(n))
                if len(nums) == 2:
                    break
            if len(nums) < 2:
                continue

            bloque = out.periodos.setdefault(periodo, {"local": {}, "cnsf": {}})
            bloque["local"][concepto.id] = nums[0]  # metodología local (para contraste)
            bloque["cnsf"][concepto.id] = nums[1]   # método estatutario CNSF
            usada = True
        if usada:
            out.hojas.append(nombre)
    return out


# =========================================================================
# 6. Histórico: cada archivo actualiza su periodo y conserva los demás
# =========================================================================


class Historico:
    def __init__(self, ruta: str | Path = "historico_reservas.json"):
        self.ruta = Path(ruta)
        self.datos: dict[str, dict[str, Any]] = {}
        if self.ruta.exists():
            self.datos = json.loads(self.ruta.read_text(encoding="utf-8"))

    # ---------------------------------------------------------------- io
    def guardar(self) -> None:
        self.ruta.write_text(json.dumps(self.datos, indent=2, ensure_ascii=False), encoding="utf-8")

    def periodos(self) -> list[str]:
        return sorted(self.datos)

    def _entrada(self, p: str) -> dict[str, Any]:
        return self.datos.setdefault(p, {"periodo": p, "local": {}, "cnsf": {}, "origen": {}})

    # ----------------------------------------------------------- ingesta
    def procesar(self, ruta: str | Path) -> list[str]:
        """Lee un Excel, detecta de cuál se trata y lo funde al histórico."""
        ruta = Path(ruta)
        hojas = leer_libro(ruta)
        avisos: list[str] = []

        bal = parse_balanza(hojas, ruta.name)
        if bal is not None:
            if not bal.periodo:
                raise ValueError(
                    f"{ruta.name}: se leyó la balanza pero no se identificó el periodo. "
                    "Renombra el archivo como Balanza_MMAAAA.xlsx."
                )
            e = self._entrada(bal.periodo)
            e["local"].update(bal.valores)
            e["origen"]["local"] = f"Balanza · {ruta.name}"
            self._revisar(e)
            avisos.append(
                f"{ruta.name} → balanza al {etiqueta_periodo(bal.periodo)} · " + " · ".join(bal.detalle)
                + (f" · SIN CUENTA {', '.join(bal.faltantes)}" if bal.faltantes else "")
            )
            return avisos

        act = parse_actuarios(hojas)
        if not act.periodos:
            raise ValueError(
                f"{ruta.name}: no se reconoció ni como balanza (falta la columna CUENTA) "
                "ni como archivo de actuarios (faltan los tres conceptos de reserva)."
            )
        for p in sorted(act.periodos):
            e = self._entrada(p)
            src = act.periodos[p]
            e["cnsf"].update(src["cnsf"])
            e["local_actuarios"] = src["local"]
            origen_local = e["origen"].get("local", "")
            if not origen_local or origen_local.startswith("Actuarios"):
                # la balanza manda; esto solo rellena periodos que aún no la tienen
                for cid, val in src["local"].items():
                    e["local"].setdefault(cid, val)
                if not origen_local:
                    e["origen"]["local"] = f"Actuarios · {ruta.name}"
            e["origen"]["cnsf"] = f"Actuarios · {ruta.name}"
            self._revisar(e)
        avisos.append(
            f"{ruta.name} → actuarios, {len(act.periodos)} periodo(s): "
            + ", ".join(etiqueta_corta(p) for p in sorted(act.periodos))
            + " · hojas: " + ", ".join(act.hojas)
        )
        avisos += [f"Revisar {etiqueta_periodo(p)}: {self.datos[p]['aviso']}"
                   for p in self.periodos() if self.datos[p].get("aviso")]
        return avisos

    @staticmethod
    def _revisar(e: dict[str, Any]) -> None:
        """Contrasta la columna local de la balanza contra la del archivo de actuarios."""
        e.pop("aviso", None)
        ref = e.get("local_actuarios")
        if not ref:
            return
        difs = [
            f"{c.label.replace('Reserva de ', '')}: balanza {fmt(e['local'][c.id])} vs actuarios {fmt(ref[c.id])}"
            for c in CONCEPTOS
            if c.id in e["local"] and c.id in ref and abs(e["local"][c.id] - ref[c.id]) > 0.5
        ]
        if difs:
            e["aviso"] = "Metodología local no coincide — " + "; ".join(difs)

    # -------------------------------------------------------------- vista
    def total(self, periodo: str, lado: str) -> float:
        return sum(self.datos[periodo][lado].get(c.id, 0.0) for c in CONCEPTOS)

    def diferencia(self, periodo: str, cid: str | None = None) -> float | None:
        e = self.datos[periodo]
        if cid is None:
            return self.total(periodo, "cnsf") - self.total(periodo, "local")
        if cid not in e["cnsf"] or cid not in e["local"]:
            return None
        return e["cnsf"][cid] - e["local"][cid]

    def mensajes(self, actual: str, previo: str | None) -> list[str]:
        """Los tres mensajes clave, redactados con las cifras del propio corte."""
        def mm(v: float) -> str:
            return f"{v / 1e6:,.2f}"

        d1 = self.diferencia(actual) or 0.0
        if previo is None:
            return [f"Al {etiqueta_periodo(actual)}, la diferencia entre metodologías asciende a USD {mm(d1)} MM.",
                    "Agrega un periodo anterior al histórico para comparar la evolución del diferencial."]

        d0 = self.diferencia(previo) or 0.0
        v_local = self.total(actual, "local") - self.total(previo, "local")
        v_cnsf = self.total(actual, "cnsf") - self.total(previo, "cnsf")
        rango = f"Entre {etiqueta_corta(previo).lower()} y {etiqueta_corta(actual).lower()}"

        motor = max(
            ((c, self.datos[actual]["local"].get(c.id, 0.0) - self.datos[previo]["local"].get(c.id, 0.0))
             for c in CONCEPTOS), key=lambda x: abs(x[1]), default=None)
        m1 = (f"{rango}, las reservas bajo QES Metodología local "
              f"{'aumentan' if v_local >= 0 else 'disminuyen'} USD {mm(abs(v_local))} MM")
        if motor and abs(motor[1]) > 5000:
            m1 += (f", principalmente por {'el incremento' if motor[1] >= 0 else 'la reducción'} "
                   f"de USD {mm(abs(motor[1]))} MM en la {motor[0].label.replace('Reserva de ', 'reserva de ')}")
        m1 += "."

        sube, baja = [], []
        for c in CONCEPTOS:
            d = self.datos[actual]["cnsf"].get(c.id, 0.0) - self.datos[previo]["cnsf"].get(c.id, 0.0)
            if abs(d) > 5000:
                (sube if d > 0 else baja).append((c, abs(d)))

        def lista(arr):
            return " y de ".join(f"USD {mm(d)} MM en {c.label.replace('Reserva de ', '').lower()}" for c, d in arr)

        m2 = (f"Bajo el Método Estatutario, las reservas totales "
              f"{'aumentan' if v_cnsf >= 0 else 'disminuyen'} USD {mm(abs(v_cnsf))} MM.")
        if sube and baja:
            m2 += (f" El incremento de {lista(sube)} fue "
                   f"{'parcialmente compensado' if v_cnsf >= 0 else 'más que compensado'} "
                   f"por la disminución de {lista(baja)}.")
        elif sube:
            m2 += f" El movimiento se concentra en el incremento de {lista(sube)}."
        elif baja:
            m2 += f" El movimiento se concentra en la disminución de {lista(baja)}."

        v = d1 - d0
        m3 = ("El Método Estatutario mantiene una posición superior." if d1 >= 0
              else "La Metodología local se mantiene por encima del Método Estatutario.")
        m3 += (f" La diferencia total por constitución pasa de USD {mm(d0)} MM en {etiqueta_corta(previo).lower()}"
               f" a USD {mm(d1)} MM en {etiqueta_corta(actual).lower()}, con "
               f"{'un incremento' if v >= 0 else 'una disminución'} de USD {mm(abs(v))} MM")
        if d0:
            m3 += f" ({'+' if v >= 0 else '−'}{abs(v / d0 * 100):.1f}%)"
        m3 += "."
        return [m1, m2, m3]

    def vista(self, periodos: Iterable[str] | None = None, escala: float = 1e6) -> str:
        """La vista comparativa en texto: conceptos por periodo más el total."""
        ps = list(periodos) if periodos else self.periodos()[-3:]
        if not ps:
            return "Histórico vacío."
        ancho = max(len(c.label) for c in CONCEPTOS) + 2
        sep = "  "

        def celda(v: float | None) -> str:
            return "—".rjust(11) if v is None else f"{v / escala:,.2f}".rjust(11)

        cab1 = "".ljust(ancho) + sep + sep.join(etiqueta_periodo(p).center(37) for p in ps)
        cab2 = "RESERVA".ljust(ancho) + sep + sep.join(
            "local".rjust(11) + sep + "CNSF".rjust(11) + sep + "dif.".rjust(11) for _ in ps)
        if len(ps) > 1:
            cab1 += sep + f"vs {etiqueta_corta(ps[-2])}".center(11)
            cab2 += sep + "incremento".rjust(11)
        lineas = [cab1, cab2, "-" * len(cab2)]

        for c in CONCEPTOS:
            fila = c.label.ljust(ancho)
            for p in ps:
                e = self.datos[p]
                fila += sep + celda(e["local"].get(c.id)) + sep + celda(e["cnsf"].get(c.id)) + sep + celda(self.diferencia(p, c.id))
            if len(ps) > 1:
                a, b = self.diferencia(ps[-1], c.id), self.diferencia(ps[-2], c.id)
                fila += sep + celda(a - b if a is not None and b is not None else None)
            lineas.append(fila)

        fila = "TOTAL RESERVAS".ljust(ancho)
        for p in ps:
            fila += sep + celda(self.total(p, "local")) + sep + celda(self.total(p, "cnsf")) + sep + celda(self.diferencia(p))
        if len(ps) > 1:
            fila += sep + celda(self.diferencia(ps[-1]) - self.diferencia(ps[-2]))
        lineas += ["-" * len(cab2), fila]

        unidad = "millones de USD" if escala == 1e6 else "USD"
        lineas.append("")
        lineas.append(f"Cifras en {unidad}. Diferencia = Método Estatutario CNSF − Metodología local.")
        return "\n".join(lineas)


# =========================================================================
# 7. Exportar la vista a un Excel
# =========================================================================


def exportar_excel(hist: "Historico", destino: str | Path,
                   periodos: Sequence[str] | None = None, escala: float = 1e6) -> Path:
    """Escribe la vista comparativa en un .xlsx con el mismo acomodo de la foto."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ps = list(periodos) if periodos else hist.periodos()[-3:]
    if not ps:
        raise ValueError("No hay periodos que exportar.")
    previo = ps[-2] if len(ps) > 1 else None

    plum = PatternFill("solid", fgColor="5B1A44")
    teal = PatternFill("solid", fgColor="134E63")
    teal2 = PatternFill("solid", fgColor="1B6C86")
    hielo = PatternFill("solid", fgColor="F5F9FB")
    blanco = Font(color="FFFFFF", bold=True)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fino = Side(style="thin", color="C9D8E1")
    borde = Border(left=fino, right=fino, top=fino, bottom=fino)
    formato = "#,##0.00"

    wb = Workbook()
    ws = wb.active
    ws.title = "Reservas QES"

    ws["A1"] = "Resultados reservas técnicas QES"
    ws["A1"].font = Font(size=16, bold=True, color="5B1A44")
    ws["A2"] = "Comparativo de metodologías y evolución del diferencial"
    ws["A2"].font = Font(size=11, color="4E6069")
    ws["A3"] = ("Cifras en millones de USD" if escala == 1e6 else "Cifras en USD") + \
               ". Diferencia = Método Estatutario CNSF − Metodología local."
    ws["A3"].font = Font(size=9, italic=True, color="7B8C95")

    fila_grupo, fila_sub, fila_ini = 5, 6, 7
    ws.cell(fila_grupo, 1, "RESERVA")
    ws.merge_cells(start_row=fila_grupo, start_column=1, end_row=fila_sub, end_column=1)

    col = 2
    for p in ps:
        ws.cell(fila_grupo, col, etiqueta_periodo(p).capitalize())
        ws.merge_cells(start_row=fila_grupo, start_column=col, end_row=fila_grupo, end_column=col + 2)
        for k, txt in enumerate(("Metodología local", "CNSF Método Estatutario", "Diferencia")):
            ws.cell(fila_sub, col + k, txt)
        col += 3
    if previo:
        ws.cell(fila_grupo, col, f"Incremento respecto de {etiqueta_corta(previo)}")
        ws.merge_cells(start_row=fila_grupo, start_column=col, end_row=fila_sub, end_column=col)

    for fila in (fila_grupo, fila_sub):
        for c in range(1, col + 1):
            celda = ws.cell(fila, c)
            celda.fill = plum if c == 1 else (teal if fila == fila_grupo else teal2)
            celda.font = blanco
            celda.alignment = centro
            celda.border = borde

    def escribe(fila: int, valor: float | None, columna: int, negrita: bool = False) -> None:
        celda = ws.cell(fila, columna, None if valor is None else valor / escala)
        celda.number_format = formato
        celda.border = borde
        if negrita:
            celda.fill = teal
            celda.font = blanco

    fila = fila_ini
    for i, c in enumerate(CONCEPTOS):
        ws.cell(fila, 1, c.label).border = borde
        col = 2
        for p in ps:
            escribe(fila, hist.datos[p]["local"].get(c.id), col)
            escribe(fila, hist.datos[p]["cnsf"].get(c.id), col + 1)
            escribe(fila, hist.diferencia(p, c.id), col + 2)
            col += 3
        if previo:
            a, b = hist.diferencia(ps[-1], c.id), hist.diferencia(previo, c.id)
            escribe(fila, a - b if a is not None and b is not None else None, col)
        if i % 2:
            for cc in range(1, col + 1):
                ws.cell(fila, cc).fill = hielo
        fila += 1

    ws.cell(fila, 1, "TOTAL RESERVAS").font = blanco
    ws.cell(fila, 1).fill = teal
    ws.cell(fila, 1).border = borde
    col = 2
    for p in ps:
        escribe(fila, hist.total(p, "local"), col, True)
        escribe(fila, hist.total(p, "cnsf"), col + 1, True)
        escribe(fila, hist.diferencia(p), col + 2, True)
        col += 3
    if previo:
        escribe(fila, (hist.diferencia(ps[-1]) or 0) - (hist.diferencia(previo) or 0), col, True)

    fila += 2
    for m in hist.mensajes(ps[-1], previo):
        ws.cell(fila, 1, "• " + m).font = Font(size=9, color="16252D")
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=max(col, 4))
        fila += 1

    ws.column_dimensions["A"].width = 38
    for i in range(2, col + 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 15
    ws.row_dimensions[fila_sub].height = 32
    ws.freeze_panes = ws.cell(fila_ini, 2)

    destino = Path(destino)
    wb.save(destino)
    return destino


# =========================================================================
# 8. Ventana: arrastrar los Excel y ver la vista
# =========================================================================


def abrir_gui(hist_ruta: str | Path = "historico_reservas.json", bloquear: bool = True):
    """Abre la ventana. Arrastra ahí los dos Excel y la vista se arma sola.

    El arrastrar y soltar necesita tkinterdnd2 (pip install tkinterdnd2); sin él
    la ventana funciona igual con el botón de elegir archivos.

    En Jupyter, ejecuta antes «%gui tk» en una celda y llama
    abrir_gui(..., bloquear=False): así la ventana vive sin dejar la celda
    ocupada. Devuelve la ventana, por si quieres cerrarla con .destroy().
    """
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    try:
        try:
            from tkinterdnd2 import DND_FILES, TkinterDnD
            root = TkinterDnD.Tk()
            arrastre = True
        except ImportError:
            root = tk.Tk()
            arrastre = False
    except tk.TclError as err:
        raise SystemExit(
            "No se pudo abrir una ventana: este Python no tiene pantalla a la mano "
            f"({err}).\n"
            "Esto suele ocurrir cuando el kernel de Jupyter corre en un servidor, en WSL "
            "o en un contenedor: ahí no hay escritorio donde dibujar.\n"
            "Opciones: correr «python reservas_qes.py» desde el Anaconda Prompt de tu "
            "propia máquina, usar Historico(...).vista() sin ventana, o abrir la versión "
            "web (reservas/index.html) en el navegador."
        ) from err

    PLUM, TEAL, TEAL2 = "#5B1A44", "#134E63", "#1B6C86"
    HIELO, HIELO2, LINEA = "#E9F1F6", "#F5F9FB", "#C9D8E1"
    TINTA, TINTA2, TINTA3 = "#16252D", "#4E6069", "#7B8C95"
    FONDO, PAPEL = "#EFF3F6", "#FFFFFF"

    import tkinter.font as tkfont
    familias = set(tkfont.families(root))
    UI = "Segoe UI" if "Segoe UI" in familias else ("Helvetica" if "Helvetica" in familias else "TkDefaultFont")
    MONO = "Consolas" if "Consolas" in familias else "TkFixedFont"

    root.title("Reservas técnicas QES")
    ancho = min(1320, root.winfo_screenwidth() - 60)
    alto = min(900, root.winfo_screenheight() - 90)
    root.geometry(f"{ancho}x{alto}+{max(0, (root.winfo_screenwidth() - ancho) // 2)}+20")
    root.configure(bg=FONDO)
    root.minsize(760, 520)

    hist = Historico(hist_ruta)
    estado = {"seleccion": hist.periodos()[-3:], "escala": 1e6, "fx": 17.4986}

    lienzo = tk.Canvas(root, bg=FONDO, highlightthickness=0)
    barra = ttk.Scrollbar(root, orient="vertical", command=lienzo.yview)
    cuerpo = tk.Frame(lienzo, bg=FONDO)
    cuerpo.bind("<Configure>", lambda e: lienzo.configure(scrollregion=lienzo.bbox("all")))
    ventana_id = lienzo.create_window((0, 0), window=cuerpo, anchor="nw")
    lienzo.bind("<Configure>", lambda e: lienzo.itemconfigure(ventana_id, width=e.width))
    lienzo.configure(yscrollcommand=barra.set)
    lienzo.pack(side="left", fill="both", expand=True)
    barra.pack(side="right", fill="y")
    root.bind_all("<MouseWheel>", lambda e: lienzo.yview_scroll(int(-e.delta / 120), "units"))
    root.bind_all("<Button-4>", lambda e: lienzo.yview_scroll(-1, "units"))
    root.bind_all("<Button-5>", lambda e: lienzo.yview_scroll(1, "units"))

    def tarjeta(padre) -> tk.Frame:
        marco = tk.Frame(padre, bg=PAPEL, highlightbackground=LINEA, highlightthickness=1)
        marco.pack(fill="x", padx=16, pady=8)
        return marco

    # ------------------------------------------------------------ encabezado
    cab = tk.Frame(cuerpo, bg=FONDO)
    cab.pack(fill="x", padx=16, pady=(14, 0))
    tk.Label(cab, text="RESERVAS TÉCNICAS QES", bg=FONDO, fg=PLUM,
             font=(UI, 17, "bold")).pack(anchor="w")
    tk.Label(cab, text="Metodología local (balanza contable) contra Método Estatutario CNSF, "
                       "con histórico acumulado por periodo.",
             bg=FONDO, fg=TINTA2, font=(UI, 9)).pack(anchor="w")

    # ------------------------------------------------------------ zona de carga
    carga = tarjeta(cuerpo)
    zona = tk.Label(
        carga,
        text=("Arrastra aquí la balanza (.xlsx) y el archivo de actuarios (.xlsb)"
              if arrastre else "Elige la balanza (.xlsx) y el archivo de actuarios (.xlsb)"),
        bg=HIELO2, fg=TEAL, font=(UI, 11, "bold"), height=2, relief="ridge", bd=1, cursor="hand2")
    zona.pack(fill="x", padx=14, pady=(14, 6))

    pie_zona = tk.Frame(carga, bg=PAPEL)
    pie_zona.pack(fill="x", padx=14, pady=(0, 10))
    tk.Label(pie_zona, text="El origen de cada archivo se detecta solo · "
                            "2205 riesgos en curso · 2301 siniestros reportados · 2302 no reportados (3er grado)",
             bg=PAPEL, fg=TINTA3, font=(UI, 8)).pack(side="left")

    bitacora = tk.Text(carga, height=3, bg=PAPEL, fg=TINTA2, font=(MONO, 8),
                       relief="flat", wrap="word", highlightbackground=LINEA, highlightthickness=1)
    bitacora.pack(fill="x", padx=14, pady=(0, 14))
    bitacora.tag_configure("err", foreground="#B04234")
    bitacora.tag_configure("warn", foreground="#9C6A0E")
    bitacora.insert("end", "Sin archivos cargados en esta sesión.\n")
    bitacora.configure(state="disabled")

    def apunta(texto: str, tag: str = "") -> None:  # noqa: D401
        bitacora.configure(state="normal")
        if bitacora.get("1.0", "end").strip() == "Sin archivos cargados en esta sesión.":
            bitacora.delete("1.0", "end")
        bitacora.insert("end", texto + "\n", tag)
        bitacora.see("end")
        bitacora.configure(state="disabled")

    def reporta_error(tipo, valor, rastro) -> None:
        import traceback
        apunta("! " + "".join(traceback.format_exception_only(tipo, valor)).strip(), "err")
        traceback.print_exception(tipo, valor, rastro)

    root.report_callback_exception = reporta_error

    # ------------------------------------------------------------ controles
    ctl = tarjeta(cuerpo)
    fila_ctl = tk.Frame(ctl, bg=PAPEL)
    fila_ctl.pack(fill="x", padx=14, pady=12)
    tk.Label(fila_ctl, text="PERIODOS EN LA VISTA", bg=PAPEL, fg=TINTA3,
             font=(UI, 8, "bold")).grid(row=0, column=0, sticky="w")
    chips = tk.Frame(fila_ctl, bg=PAPEL)
    chips.grid(row=1, column=0, sticky="w", pady=(4, 0))

    tk.Label(fila_ctl, text="UNIDADES", bg=PAPEL, fg=TINTA3,
             font=(UI, 8, "bold")).grid(row=0, column=1, sticky="w", padx=(28, 0))
    unidad_var = tk.StringVar(value="Millones de USD")
    combo = ttk.Combobox(fila_ctl, textvariable=unidad_var, state="readonly", width=18,
                         values=("Millones de USD", "USD", "Millones de MXN"))
    combo.grid(row=1, column=1, sticky="w", padx=(28, 0), pady=(4, 0))

    tk.Label(fila_ctl, text="TIPO DE CAMBIO MXN/USD", bg=PAPEL, fg=TINTA3,
             font=(UI, 8, "bold")).grid(row=0, column=2, sticky="w", padx=(20, 0))
    fx_var = tk.StringVar(value="17.4986")
    tk.Entry(fila_ctl, textvariable=fx_var, width=10, font=(MONO, 9)).grid(
        row=1, column=2, sticky="w", padx=(20, 0), pady=(4, 0))

    botones = tk.Frame(fila_ctl, bg=PAPEL)
    botones.grid(row=1, column=3, sticky="e", padx=(24, 0))
    fila_ctl.columnconfigure(3, weight=1)

    # ------------------------------------------------------------ tablero
    kpis = tarjeta(cuerpo)
    tabla = tarjeta(cuerpo)
    claves = tarjeta(cuerpo)

    def limpia(marco: tk.Frame) -> None:
        for hijo in marco.winfo_children():
            hijo.destroy()

    def escala_valor(v: float) -> float:
        if estado["escala"] == 1.0:
            return v
        if estado["escala"] == "mxn":
            return v * estado["fx"] / 1e6
        return v / 1e6

    def texto_valor(v: float | None) -> str:
        if v is None:
            return "n/d"
        e = escala_valor(v)
        return "—" if abs(e) < 0.005 else f"{e:,.2f}"

    def mm(v: float) -> str:
        return f"{v / 1e6:,.2f}"

    # ------------------------------------------------------------ pintado
    def dibuja() -> None:
        ps = sorted(estado["seleccion"])
        actual = ps[-1] if ps else None
        previo = ps[-2] if len(ps) > 1 else None

        # chips de periodos
        limpia(chips)
        if not hist.periodos():
            tk.Label(chips, text="Sin periodos en el histórico", bg=PAPEL, fg=TINTA3,
                     font=(UI, 9)).pack(side="left")
        for p in hist.periodos():
            var = tk.BooleanVar(value=p in ps)

            def alterna(p=p, var=var):
                if var.get():
                    estado["seleccion"] = sorted(set(estado["seleccion"]) | {p})
                else:
                    estado["seleccion"] = [x for x in estado["seleccion"] if x != p] or [p]
                dibuja()

            tk.Checkbutton(chips, text=etiqueta_periodo(p), variable=var, command=alterna,
                           bg=PAPEL, fg=TINTA, activebackground=PAPEL, font=(UI, 9),
                           selectcolor=PAPEL).pack(side="left", padx=(0, 10))

        # indicadores
        limpia(kpis)
        caja = tk.Frame(kpis, bg=PAPEL)
        caja.pack(fill="x", padx=14, pady=12)
        if actual is None:
            tk.Label(caja, text="Carga un archivo para ver los indicadores.",
                     bg=PAPEL, fg=TINTA2, font=(UI, 10)).pack(anchor="w")
        else:
            d1 = hist.diferencia(actual) or 0.0
            tarjetas = [("DIFERENCIA TOTAL", etiqueta_corta(actual),
                         f"USD {mm(d1)} MM", f"~MXN {mm(d1 * estado['fx'])} MM", TEAL, HIELO2)]
            if previo:
                d0 = hist.diferencia(previo) or 0.0
                v = d1 - d0
                tarjetas.append(("VARIACIÓN DEL PERIODO", f"{etiqueta_corta(previo)} → {etiqueta_corta(actual)}",
                                 f"{'+' if v >= 0 else '−'}USD {mm(abs(v))} MM",
                                 f"~MXN {mm(abs(v) * estado['fx'])} MM", PLUM, "#FBF1F7"))
                pct = f"{'+' if v >= 0 else '−'}{abs(v / d0 * 100):.1f}%" if d0 else "n/d"
                tarjetas.append(("VARIACIÓN % DE LA DIFERENCIA",
                                 f"{etiqueta_corta(previo)} → {etiqueta_corta(actual)}",
                                 pct, f"sobre USD {mm(d0)} MM", TEAL, HIELO2))
            for i, (etq, alcance, valor, alt, color, fondo) in enumerate(tarjetas):
                t = tk.Frame(caja, bg=fondo, highlightbackground=LINEA, highlightthickness=1)
                t.grid(row=0, column=i, sticky="ew", padx=(0 if i == 0 else 10, 0))
                caja.columnconfigure(i, weight=1)
                tk.Label(t, text=etq, bg=fondo, fg=color, font=(UI, 8, "bold")).pack(anchor="w", padx=12, pady=(8, 0))
                tk.Label(t, text=alcance, bg=fondo, fg=TINTA3, font=(UI, 8)).pack(anchor="w", padx=12)
                tk.Label(t, text=valor, bg=fondo, fg=color, font=(UI, 19, "bold")).pack(anchor="w", padx=12)
                tk.Label(t, text=alt, bg=fondo, fg=TINTA3, font=(MONO, 8)).pack(anchor="w", padx=12, pady=(0, 8))

        # matriz
        limpia(tabla)
        envoltura = tk.Frame(tabla, bg=PAPEL)
        envoltura.pack(fill="x", padx=14, pady=(14, 6))
        lienzo_t = tk.Canvas(envoltura, bg=PAPEL, highlightthickness=0)
        barra_h = ttk.Scrollbar(envoltura, orient="horizontal", command=lienzo_t.xview)
        lienzo_t.configure(xscrollcommand=barra_h.set)
        lienzo_t.pack(side="top", fill="x")
        barra_h.pack(side="bottom", fill="x")
        rejilla = tk.Frame(lienzo_t, bg=LINEA)
        lienzo_t.create_window((0, 0), window=rejilla, anchor="nw")

        def ajusta(_evento=None):
            lienzo_t.configure(scrollregion=lienzo_t.bbox("all"),
                               height=rejilla.winfo_reqheight())

        rejilla.bind("<Configure>", ajusta)
        if not ps:
            tk.Label(rejilla, text="Selecciona al menos un periodo.", bg=PAPEL, fg=TINTA2,
                     font=(UI, 10)).grid(row=0, column=0, sticky="w")
        else:
            def celda(r, c, texto, **kw):
                lab = tk.Label(rejilla, text=texto, bg=kw.get("bg", PAPEL), fg=kw.get("fg", TINTA),
                               font=kw.get("font", (UI, 9)), anchor=kw.get("anchor", "e"),
                               padx=8, pady=5, wraplength=kw.get("wrap", 0), justify="center")
                lab.grid(row=r, column=c, sticky="nsew", padx=(0, 1), pady=(0, 1),
                         columnspan=kw.get("cs", 1), rowspan=kw.get("rs", 1))
                return lab

            celda(0, 0, "RESERVA", bg=PLUM, fg="white", font=(UI, 11, "bold"), anchor="w", rs=2)
            col = 1
            for p in ps:
                celda(0, col, etiqueta_periodo(p).capitalize(), bg=TEAL, fg="white",
                      font=(UI, 10, "bold"), anchor="center", cs=3)
                for k, txt in enumerate(("Metodología\nlocal", "CNSF\nMétodo Estatutario", "Diferencia\n(estat. − local)")):
                    celda(1, col + k, txt, bg=TEAL2, fg="white", font=(UI, 8), anchor="center", wrap=110)
                col += 3
            if previo:
                celda(0, col, f"Incremento\nrespecto de\n{etiqueta_corta(previo)}", bg=TEAL, fg="white",
                      font=(UI, 8, "bold"), anchor="center", rs=2, wrap=100)

            for i, c in enumerate(CONCEPTOS):
                fondo = PAPEL if i % 2 == 0 else HIELO2
                r = 2 + i
                celda(r, 0, c.label, bg=fondo, anchor="w")
                col = 1
                for p in ps:
                    celda(r, col, texto_valor(hist.datos[p]["local"].get(c.id)), bg=fondo, font=(MONO, 9))
                    celda(r, col + 1, texto_valor(hist.datos[p]["cnsf"].get(c.id)), bg=fondo, font=(MONO, 9))
                    celda(r, col + 2, texto_valor(hist.diferencia(p, c.id)), bg=fondo, fg=PLUM, font=(MONO, 9))
                    col += 3
                if previo:
                    a, b = hist.diferencia(ps[-1], c.id), hist.diferencia(previo, c.id)
                    celda(r, col, texto_valor(a - b if a is not None and b is not None else None),
                          bg=fondo, fg=TEAL, font=(MONO, 9))

            r = 2 + len(CONCEPTOS)
            celda(r, 0, "TOTAL RESERVAS", bg=TEAL, fg="white", font=(UI, 10, "bold"), anchor="w")
            col = 1
            for p in ps:
                celda(r, col, texto_valor(hist.total(p, "local")), bg=TEAL, fg="white", font=(MONO, 9, "bold"))
                celda(r, col + 1, texto_valor(hist.total(p, "cnsf")), bg=TEAL, fg="white", font=(MONO, 9, "bold"))
                celda(r, col + 2, texto_valor(hist.diferencia(p)), bg=TEAL, fg="white", font=(MONO, 9, "bold"))
                col += 3
            if previo:
                celda(r, col, texto_valor((hist.diferencia(ps[-1]) or 0) - (hist.diferencia(previo) or 0)),
                      bg=TEAL, fg="white", font=(MONO, 9, "bold"))
            rejilla.columnconfigure(0, minsize=250)
            for c in range(1, col + 1):
                rejilla.columnconfigure(c, minsize=96)
        ajusta()

        unidad = {1.0: "USD", "mxn": "millones de MXN"}.get(estado["escala"], "millones de USD")
        tk.Label(tabla, text=f"Cifras en {unidad}. La diferencia es Método Estatutario CNSF menos Metodología "
                             "local, es decir el exceso de constitución del método estatutario.",
                 bg=PAPEL, fg=TINTA2, font=(UI, 8), anchor="w", justify="left").pack(
            fill="x", padx=14, pady=(0, 12))

        # mensajes clave
        limpia(claves)
        tk.Label(claves, text="MENSAJES CLAVE", bg=PAPEL, fg=TEAL2, font=(UI, 8, "bold")).pack(
            anchor="w", padx=14, pady=(12, 4))
        if actual is None:
            tk.Label(claves, text="Carga los dos archivos para redactar los mensajes.",
                     bg=PAPEL, fg=TINTA2, font=(UI, 9)).pack(anchor="w", padx=14, pady=(0, 12))
        else:
            for m in hist.mensajes(actual, previo):
                tk.Label(claves, text="•  " + m, bg=PAPEL, fg=TINTA, font=(UI, 9),
                         wraplength=1120, justify="left", anchor="w").pack(fill="x", padx=14, pady=(0, 6))
            tk.Label(claves, text="", bg=PAPEL).pack(pady=(0, 6))

    # ------------------------------------------------------------ acciones
    def carga_rutas(rutas: Iterable[str]) -> None:
        for ruta in rutas:
            if not str(ruta).strip():
                continue
            try:
                for aviso in hist.procesar(ruta):
                    apunta("· " + aviso, "warn" if "Revisar" in aviso or "SIN CUENTA" in aviso else "")
            except Exception as err:  # noqa: BLE001 - se reporta en la bitácora
                apunta("! " + str(err), "err")
                continue
            nuevos = [p for p in hist.periodos() if p not in estado["seleccion"]]
            if nuevos:
                estado["seleccion"] = sorted(set(estado["seleccion"]) | {nuevos[-1]})[-4:]
        if len(estado["seleccion"]) < 3:      # arranca mostrando los tres últimos cortes
            estado["seleccion"] = sorted(set(estado["seleccion"]) | set(hist.periodos()[-3:]))[-4:]
        hist.guardar()
        dibuja()

    def elegir() -> None:
        rutas = filedialog.askopenfilenames(
            title="Elige la balanza y el archivo de actuarios",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xlsb"), ("Todos", "*.*")])
        if rutas:
            carga_rutas(rutas)

    zona.bind("<Button-1>", lambda e: elegir())
    if arrastre:
        zona.drop_target_register(DND_FILES)
        zona.dnd_bind("<<Drop>>", lambda e: carga_rutas(root.tk.splitlist(e.data)))
        zona.dnd_bind("<<DragEnter>>", lambda e: zona.configure(bg=HIELO))
        zona.dnd_bind("<<DragLeave>>", lambda e: zona.configure(bg=HIELO2))
    else:
        tk.Label(pie_zona, text="  (instala tkinterdnd2 para arrastrar y soltar)",
                 bg=PAPEL, fg="#9C6A0E", font=(UI, 8)).pack(side="left")

    def cambia_unidad(_evento=None) -> None:
        estado["escala"] = {"Millones de USD": 1e6, "USD": 1.0, "Millones de MXN": "mxn"}[unidad_var.get()]
        dibuja()

    def cambia_fx(*_args) -> None:
        try:
            v = float(fx_var.get().replace(",", ""))
            if v > 0:
                estado["fx"] = v
                dibuja()
        except ValueError:
            pass

    combo.bind("<<ComboboxSelected>>", cambia_unidad)
    fx_var.trace_add("write", cambia_fx)

    def exporta() -> None:
        if not estado["seleccion"]:
            messagebox.showinfo("Reservas QES", "No hay periodos que exportar.")
            return
        destino = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile="Vista_reservas_QES.xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not destino:
            return
        try:
            ruta = exportar_excel(hist, destino, sorted(estado["seleccion"]),
                                  escala=1e6 if estado["escala"] == "mxn" else estado["escala"])
            apunta(f"· vista exportada a {ruta}")
            messagebox.showinfo("Reservas QES", f"Vista exportada:\n{ruta}")
        except Exception as err:  # noqa: BLE001
            messagebox.showerror("Reservas QES", str(err))

    def vacia() -> None:
        if messagebox.askyesno("Reservas QES", "¿Vaciar todo el histórico guardado?"):
            hist.datos.clear()
            estado["seleccion"] = []
            hist.guardar()
            apunta("· histórico vaciado")
            dibuja()

    ttk.Button(botones, text="Elegir archivos", command=elegir).pack(side="left", padx=(0, 6))
    ttk.Button(botones, text="Exportar vista a Excel", command=exporta).pack(side="left", padx=(0, 6))
    ttk.Button(botones, text="Vaciar histórico", command=vacia).pack(side="left")

    dibuja()
    apunta(f"· histórico: {hist.ruta} ({len(hist.periodos())} periodo(s))")

    # que no nazca detrás del navegador ni del notebook
    try:
        root.update_idletasks()
        root.deiconify()
        root.lift()
        root.attributes("-topmost", True)
        root.after(800, lambda: root.attributes("-topmost", False))
        root.focus_force()
    except tk.TclError:
        pass

    print(f"Ventana abierta ({ancho}x{alto}). Ciérrala para liberar la celda.")

    if bloquear:
        root.mainloop()
    else:
        root.update()          # con «%gui tk» IPython se encarga del resto
    return root


# =========================================================================
# 9. Línea de comandos
# =========================================================================


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Arma la vista de reservas técnicas QES desde los dos Excel.")
    ap.add_argument("archivos", nargs="*", help="Balanza (.xlsx) y archivo de actuarios (.xlsb/.xlsx), en cualquier orden.")
    ap.add_argument("--hist", default="historico_reservas.json", help="JSON donde vive el histórico acumulado.")
    ap.add_argument("--periodos", nargs="*", help="Periodos a mostrar (AAAA-MM-DD). Por omisión, los tres últimos.")
    ap.add_argument("--usd", action="store_true", help="Mostrar en USD en vez de millones.")
    ap.add_argument("--gui", action="store_true", help="Abrir la ventana (es lo que pasa si no das archivos).")
    ap.add_argument("--excel", help="Escribir la vista en este .xlsx además de imprimirla.")
    args = ap.parse_args(argv)

    if args.gui or not args.archivos:
        abrir_gui(args.hist)
        return 0

    hist = Historico(args.hist)
    for ruta in args.archivos:
        try:
            for aviso in hist.procesar(ruta):
                print(f"· {aviso}", file=sys.stderr)
        except (ValueError, SystemExit) as err:
            print(f"! {err}", file=sys.stderr)
            return 1
    if args.archivos:
        hist.guardar()
        print(f"· histórico guardado en {hist.ruta} ({len(hist.periodos())} periodo(s))\n", file=sys.stderr)

    print(hist.vista(args.periodos, escala=1.0 if args.usd else 1e6))
    if args.excel:
        print(f"· vista exportada a {exportar_excel(hist, args.excel, args.periodos)}", file=sys.stderr)
    return 0


if "ipykernel" in sys.modules:
    # dentro de Jupyter no se arranca la CLI: ahí argparse vería los argumentos
    # del kernel ("-f <connection file>") y abortaría. La ventana se pide a mano.
    print("Listo. Para abrir la ventana ejecuta en la siguiente celda:\n"
          "    abrir_gui(r\"historico_reservas.json\")\n"
          "Sin ventana: h = Historico(\"historico_reservas.json\"); "
          "h.procesar(ruta); h.guardar(); print(h.vista())")
elif __name__ == "__main__":
    raise SystemExit(main())
